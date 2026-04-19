from flask import Flask, request, jsonify
from openpyxl import load_workbook
import io, re, base64

app = Flask(__name__)

# Account row label → output field name
ACCOUNT_MAP = {
    # Income
    "4000 Services":                                "income_services",
    "4050 Scrap Metal Sales":                       "income_scrap_metal",
    "4080 Equipter Lease Income":                   "income_equipter",
    "4090 Returns":                                 "income_returns",
    "Job Income":                                   "income_job",
    "Unapplied Cash Payment Income":                "unapplied_cash",
    "Total for Income":                             "income_total",
    # COGS
    "5010 Business licenses & permits":             "licenses_permits",
    "5020 Commissions COGS":                        "commissions_cogs",
    "5030 Cost of goods sold":                      "cogs_base",
    "6115 WCP -Builders Mutual Policy":             "wcp_builders_mutual",
    "Bond":                                         "bond",
    "Total for 5030 Cost of goods sold":            "cogs_30_total",
    "5035 Fortified Inspections":                   "fortified_inspections",
    "5040 Equipment Rental for Jobs":               "equipment_rental",
    "Equipment Hauling":                            "equipment_hauling",
    "Total for 5040 Equipment Rental for Jobs":     "equipment_total",
    "5200 Fuel & Gas":                              "fuel_gas",
    "5210 Hotels/Travel COGS":                      "hotels_travel",
    "5220 Job Plans":                               "job_plans",
    "5230 Job Supplies & Materials":                "materials",
    "5240 Management Fees COGS":                    "mgmt_fees_cogs",
    "5250 Other Miscellaneous Service Cost":        "misc_service_cost",
    "5260 Permits":                                 "permits",
    "5270 Purchases":                               "purchases",
    "5280 Shipping":                                "shipping",
    "5290 Subcontractor expenses":                  "subcontractors",
    "5295 Labor Service Fees":                      "labor_service_fees",
    "Total for 5290 Subcontractor expenses":        "subcontractors_total",
    "5300 Tool Inventory":                          "tool_inventory",
    "5310 Waste Removal":                           "waste_removal",
    "5320 Commercial Package Policy w/BM":          "commercial_package",
    "5100 Field Payroll Taxes":                     "field_payroll_taxes",
    "5105 Field PTO":                               "field_pto",
    "5110 Field Salaries & Wages":                  "field_wages",
    "5115 Field Staff 401K Match":                  "field_401k",
    "5120 Field Staff Health Insurance":            "field_health_insurance",
    "5125 Field Staff Payroll Fees":                "field_payroll_fees",
    "Total for Field Staff Payroll":                "labor_total",
    "Total for Cost of Goods Sold":                 "total_cogs",
    # Key summary rows
    "Gross Profit":                                 "gross_profit",
    # Office expense sub-items
    "6200 Building Security":                       "office_building_security",
    "6205 Merchant account fees":                   "office_merchant_fees",
    "6210 Office Cleaning":                         "office_cleaning",
    "6215 Office Devices & Repairs":                "office_devices",
    "6220 Office Food & Beverage":                  "office_food",
    "6225 Office supplies":                         "office_supplies",
    "6230 Shipping & postage":                      "office_shipping",
    "6235 Software & apps":                         "office_software",
    # Payroll sub-items
    "6305 Employee retirement plans":               "payroll_retirement",
    "Combined Insurance":                           "payroll_combined_insurance",
    "6355 Overhead Payroll Taxes":                  "overhead_payroll_taxes",
    "6385 Production Payroll Taxes":                "production_payroll_taxes",
    "6395 Production Salaries & Wages":             "production_wages",
    "6425 Sales Staff Payroll Taxes":               "sales_payroll_taxes",
    "6430 Sales Staff Salaries & Wages":            "sales_wages",
    "6445 Sheet Metal Shop Payroll Taxes":          "shop_payroll_taxes",
    "6450 Sheet Metal Shop Salaries & Wages":       "shop_wages",
    "Shop Health Insurance":                        "shop_health_insurance",
    # Utilities sub-items
    "6660 Phone service":                           "utilities_phone",
    "Internet & TV services":                       "utilities_internet",
    # Vehicle sub-items
    "6675 Auto Insurance CAP Policy w/BM":          "vehicle_auto_insurance",
    "6680 Parking & Tolls":                         "vehicle_parking",
    "6685 Vehicle Registration":                    "vehicle_registration",
    # Other income
    "7000 Interest Earned":                         "other_interest",
    "7110 Credit card rewards":                     "other_credit_rewards",
    "7120 Rebates":                                 "other_rebates",
    "7130 Sales Tax Discount":                      "other_sales_tax_discount",
    "7140 Unrealized Gain (Loss)":                  "other_unrealized_gain",
    "Total for Other Income":                       "other_income_total",
    # Other expenses
    "8000 Ask My Accountant":                       "other_exp_ask_accountant",
    "8200 Depreciation expense":                    "other_exp_depreciation",
    "8500 Reconciliation Discrepancies":            "other_exp_reconciliation",
    "8505 Penalties & Interest":                    "other_exp_penalties",
    "Total for Other Expenses":                     "other_expenses_total",
    "Net Other Income":                             "net_other_income",
    # Expenses
    "6010 Advertising & marketing":                 "advertising",
    "6020 Auto & truck expense":                    "auto_truck",
    "6030 Bank fees & service charges":             "bank_fees",
    "6040 Building & property rent":                "building_rent",
    "6050 Continuing education":                    "continuing_education",
    "6060 Contract labor":                          "contract_labor",
    "6070 Employee Drug Testing":                   "drug_testing",
    "6080 General business expenses":               "general_business",
    "6085 Bad Debt":                                "bad_debt",
    "Total for 6080 General business expenses":     "general_business_total",
    "6100 Insurance":                               "insurance",
    "6105 MUB Umbrella General Liability Insurance":"umbrella_insurance",
    "6110 Property insurance":                      "property_insurance",
    "Total for 6100 Insurance":                     "insurance_total",
    "6120 Interest expense":                        "interest_expense",
    "6130 Legal & accounting services":             "legal_accounting",
    "6135 Accounting fees":                         "accounting_fees",
    "6140 Legal fees":                              "legal_fees",
    "Total for 6130 Legal & accounting services":   "legal_accounting_total",
    "6150 Meals & Entertainment":                   "meals_entertainment",
    "6170 Memberships":                             "memberships",
    "6180 Monthly Subscriptions":                   "subscriptions",
    "6190 Office expenses":                         "office_expenses",
    "Total for 6190 Office expenses":               "office_expenses_total",
    "6300 Payroll expenses":                        "payroll_expenses",
    "6325 Group term life insurance":               "group_life_insurance",
    "6330 Health & accident plans":                 "health_accident",
    "6400 Salaries & Wages (clearing)":             "salaries_wages_clearing",
    "6490 Uniforms":                                "uniforms",
    "6500 Virtual Assistants":                      "virtual_assistants",
    "6510 Vision and Dental Insurance":             "vision_dental",
    # Overhead Payroll
    "6340 Overhead 401K Match":                     "overhead_401k",
    "6345 Overhead Health Insurance":               "overhead_health",
    "6350 Overhead Payroll Fees":                   "overhead_payroll_fees",
    "6355 Overhead Payroll Taxes":                  "overhead_payroll_taxes",
    "6360 Overhead Salaries & Wages":               "overhead_wages",
    "Total for Overhead Payroll":                   "overhead_payroll_total",
    # Production Payroll
    "6370 Production 401K Match":                   "production_401k",
    "6375 Production Health Insurance":             "production_health",
    "6380 Production Payroll Fees":                 "production_payroll_fees",
    "6385 Production Payroll Taxes":                "production_payroll_taxes",
    "6390 Production PTO":                          "production_pto",
    "6395 Production Salaries & Wages":             "production_wages",
    "Total for Production Payroll":                 "production_payroll_total",
    # Sales Staff Payroll
    "6410 Sales Staff 401K Match":                  "sales_401k",
    "6415 Sales Staff Health Insurance":            "sales_health",
    "6420 Sales Staff Payroll Fees":                "sales_payroll_fees",
    "6425 Sales Staff Payroll Taxes":               "sales_payroll_taxes",
    "6430 Sales Staff Salaries & Wages":            "sales_wages",
    "Total for Sales Staff Payroll":                "sales_payroll_total",
    # Sheet Metal Shop Payroll
    "6440 Sheet Metal PTO":                         "shop_pto",
    "6445 Sheet Metal Shop Payroll Taxes":          "shop_payroll_taxes",
    "6450 Sheet Metal Shop Salaries & Wages":       "shop_wages",
    "6455 Shop 401K Match":                         "shop_401k",
    "6465 Shop Payroll Fees":                       "shop_payroll_fees",
    "Total for Sheet Metal Shop Payroll":           "shop_payroll_total",
    "Total for 6300 Payroll expenses":              "payroll_expenses_total",
    # Other expenses
    "6520 Professional Fees":                       "professional_fees",
    "6530 Property Maintenance":                    "property_maintenance",
    "6540 Recruiting":                              "recruiting",
    "6550 Repairs & maintenance":                   "repairs_maintenance",
    "6570 Sales Tax Expense":                       "sales_tax",
    "6580 Contractor's Tax":                        "contractors_tax",
    "Total for 6570 Sales Tax Expense":             "sales_tax_total",
    "6590 Shipping, freight & delivery":            "shipping_freight",
    "6600 Supplies & materials":                    "supplies_materials",
    "6610 Taxes paid":                              "taxes_paid",
    "6640 Utilities":                               "utilities",
    "Total for 6640 Utilities":                     "utilities_total",
    "6670 Vehicle Expense":                         "vehicle_expense",
    "Total for 6670 Vehicle Expense":               "vehicle_expense_total",
    "6691 Contributions to charities":              "charitable_contributions",
    "6696 Reimbursements":                          "reimbursements",
    "Total for Expenses":                           "total_expenses",
    "Net Operating Income":                         "net_operating_income",
    "Net Income":                                   "net_income",
}

ALL_FIELDS = sorted(ACCOUNT_MAP.values())

def clean(value):
    if value is None: return 0.0
    if isinstance(value, (int, float)): return round(float(value), 2)
    return 0.0

def parse_job_number(name):
    m = re.search(r'[-#]\s*(\d{3,5})\b', str(name))
    return m.group(1) if m else ""

def process_workbook(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Row 3 = period, Row 5 = headers (0-indexed: rows 2 and 4)
    period = str(all_rows[2][0]).strip()
    headers = all_rows[4]

    # Build row index by account name
    row_index = {}
    for i, row in enumerate(all_rows):
        label = row[0]
        if label and str(label).strip() in ACCOUNT_MAP:
            key = ACCOUNT_MAP[str(label).strip()]
            row_index[key] = i

    # Collect parent customer names from "Total for X" columns so we can exclude them
    parent_names = set()
    for h in headers:
        if h and str(h).strip().startswith("Total for "):
            parent_names.add(str(h).strip()[len("Total for "):])

    # Only keep leaf-level project columns — skip totals, parents, "Total", and blank
    project_cols = []
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h).strip()
        if h_str in ("", "Total", "Other"):
            continue
        if h_str.startswith("Total for "):
            continue
        if h_str in parent_names:
            continue
        project_cols.append(i)

    # Build records — one per project column that has any income data
    records = []
    income_row = row_index.get("income_total")

    for col_i in project_cols:
        # Skip columns with no income
        if income_row is not None:
            val = all_rows[income_row][col_i]
            if not val or not isinstance(val, (int, float)) or val == 0:
                continue

        project = str(headers[col_i]).strip()
        record = {
            "customer":   project,
            "job_number": parse_job_number(project),
            "period":     period,
            "project":    project,
        }
        for field in ALL_FIELDS:
            row_i = row_index.get(field)
            record[field] = clean(all_rows[row_i][col_i]) if row_i is not None else 0.0

        records.append(record)

    return records

@app.route('/headers', methods=['GET'])
def headers():
    base = ["customer", "project", "job_number", "period"]
    return jsonify(base + ALL_FIELDS)

@app.route('/process', methods=['POST'])
def process():
    include_headers = request.args.get('include_headers', 'false').lower() == 'true'

    def build_response(file_bytes):
        records = process_workbook(file_bytes)
        if include_headers:
            base = ["customer", "project", "job_number", "period"]
            header_row = {f: f for f in base + ALL_FIELDS}
            return jsonify([header_row] + records)
        return jsonify(records)

    file = request.files.get('file')
    if file:
        return build_response(file.read())

    if request.data:
        try:
            return build_response(request.data)
        except: pass

    if request.json:
        data = request.json.get('data') or request.json.get('file')
        if data:
            try:
                return build_response(base64.b64decode(data))
            except Exception as e:
                return jsonify({"error": str(e)}), 400

    return jsonify({"error": "No file received"}), 400

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
